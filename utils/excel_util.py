from openpyxl import Workbook, styles
import xlsxwriter
import sqlite3
import configparser


class ExcelUtil:
    @staticmethod
    def writeDeletedRows(fileName: str, fileExt: str) -> None:
        parser = configparser.ConfigParser()
        parser.read('config.ini')

        conn = sqlite3.connect(parser.defaults().get('databasename'))
        cursor = conn.cursor()
        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('oldtable')}")
        colNames = [description[0] for description in result.description]
        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('rowsdeletedtablename')}")

        suffix = 0
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(colNames)
        rowLimit = int(parser.defaults().get('rowlimit'))
        for idx, row in enumerate(result):
            if idx > 0 and idx % rowLimit == 0:
                workbook.save(f'{fileName}-{suffix}.{fileExt}')
                suffix += 1
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(colNames)
            worksheet.append(row)
        
        workbook.save(f'{fileName}-{suffix}.{fileExt}')
    
    @staticmethod
    def writeAddedRows(fileName: str, fileExt: str) -> None:
        parser = configparser.ConfigParser()
        parser.read('config.ini')

        conn = sqlite3.connect(parser.defaults().get('databasename'))
        cursor = conn.cursor()
        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('newtable')}")
        colNames = [description[0] for description in result.description]
        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('rowsinsertedtablename')}")

        suffix = 0
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(colNames)
        rowLimit = int(parser.defaults().get('rowlimit'))
        for idx, row in enumerate(result):
            if idx > 0 and idx % rowLimit == 0:
                workbook.save(f'{fileName}-{suffix}.{fileExt}')
                suffix += 1
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(colNames)
            worksheet.append(row)
        
        workbook.save(f'{fileName}-{suffix}.{fileExt}')

    @staticmethod
    def writeUpdatedRows(fileName: str, fileExt: str) -> None:
        parser = configparser.ConfigParser()
        parser.read('config.ini')

        conn = sqlite3.connect(parser.defaults().get('databasename'))
        cursor = conn.cursor()

        oldTableColLen = cursor.execute(f"SELECT COUNT(*) FROM pragma_table_info('{parser.defaults().get('oldtable')}')").fetchone()[0]
        newTableColLen = cursor.execute(f"SELECT COUNT(*) FROM pragma_table_info('{parser.defaults().get('newtable')}')").fetchone()[0]
        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('oldtable' if oldTableColLen > newTableColLen else 'newtable')}")
        colNames = [description[0] for description in result.description]

        result = cursor.execute(f"SELECT * FROM {parser.defaults().get('rowsupdatedtablename')}")

        suffix = 0
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(colNames)
        rowLimit = int(parser.defaults().get('rowlimit'))
        for idx, row in enumerate(result):
            if idx > 0 and idx % rowLimit == 0:
                workbook.save(f'{fileName}-{suffix}.{fileExt}')
                suffix += 1
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(colNames)
            row1 = row[:oldTableColLen]
            row2 = row[oldTableColLen:]
            worksheet.append(row1)
            worksheet.append(row2)
            worksheet.append([])
            for idx, data in enumerate(zip(row1, row2)):
                data1, data2 = data
                if data1 != data2:
                    column = worksheet[xlsxwriter.utility.xl_col_to_name(idx)]
                    lastRowCell = column[len(column) - 1]
                    secondLastRowCell = column[len(column) - 2]

                    red = styles.colors.Color(rgb='00FF0000')
                    redFill = styles.fills.PatternFill(patternType='solid', fgColor=red)
                    secondLastRowCell.fill = redFill

                    green = styles.colors.Color(rgb='0000FF00')
                    greenFill = styles.fills.PatternFill(patternType='solid', fgColor=green)
                    lastRowCell.fill = greenFill
        workbook.save(f'{fileName}-{suffix}.{fileExt}')

